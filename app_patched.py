#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Salon edition:
- QR en en-tête (image ou généré dynamiquement via QR_TARGET_URL)
- Bouton .vcf masquable (SHOW_DOWNLOAD_BUTTON)
- Scanner QR (caméra arrière par défaut)
- Export CSV + Excel (.xlsx) avec vrai tableau stylé
"""
from __future__ import annotations
import os
import io
import logging
from pathlib import Path
from typing import Optional, Dict

import streamlit as st
from dotenv import load_dotenv

from modules.contact import build_vcard_bytes
from modules.storage import Lead, Storage, StorageConfig
from modules.utils import ensure_data_dir, get_client_ip_hash, load_image_bytes
from modules.qr import parse_contact_from_qr

# WebRTC imports (optionnels)
try:
    import av  # noqa: F401
    import cv2  # noqa: F401
    from streamlit_webrtc import webrtc_streamer, WebRtcMode, VideoTransformerBase, RTCConfiguration
    QR_ENABLED = True
except Exception:
    QR_ENABLED = False

APP_NAME = "Tap-to-Contact — Salon"
RTC_CONFIGURATION = RTCConfiguration({
    'iceServers': [{'urls': ['stun:stun.l.google.com:19302']}]
})

# --- Boot ---
load_dotenv()
DATA_DIR = Path(os.getenv("DATA_DIR", "data"))
LOGS_DIR = Path(os.getenv("LOGS_DIR", "logs"))
ensure_data_dir(DATA_DIR)

# --- Identité (.env) ---
IDENTITY: Dict[str, str] = {
    "FN": os.getenv("FULL_NAME", "Votre Nom"),
    "N_LAST": os.getenv("N_LAST", "Nom"),
    "N_FIRST": os.getenv("N_FIRST", "Prénom"),
    "ORG": os.getenv("ORG", "Votre Société"),
    "TITLE": os.getenv("TITLE", "Fonction"),
    "TEL": os.getenv("TEL", "+33 6 00 00 00 00"),
    "EMAIL": os.getenv("EMAIL", "vous@example.com"),
    "URL": os.getenv("URL", "https://www.exemple.com"),
    "ADR_STREET": os.getenv("ADR_STREET", ""),
    "ADR_CITY": os.getenv("ADR_CITY", ""),
    "ADR_PC": os.getenv("ADR_PC", ""),
    "ADR_COUNTRY": os.getenv("ADR_COUNTRY", "France"),
}

PHOTO_PATH = os.getenv("PHOTO_PATH", "assets/photo.jpg")
SHOW_QR_IN_HEADER = os.getenv("SHOW_QR_IN_HEADER", "true").lower() in ("1","true","yes","on")
QR_IMAGE_PATH = os.getenv("QR_IMAGE_PATH", "assets/qr.png")
QR_TARGET_URL = os.getenv("QR_TARGET_URL", "")  # si défini, on génère le QR dynamiquement
SHOW_DOWNLOAD_BUTTON = os.getenv("SHOW_DOWNLOAD_BUTTON", "false").lower() in ("1","true","yes","on")
LOGO_PATH = os.getenv("LOGO_PATH", "assets/logo.png")
SHOW_LOGO = os.getenv("SHOW_LOGO", "true").lower() in ("1","true","yes","on")

# --- Storage ---
storage_cfg = StorageConfig(
    csv_path=DATA_DIR / "leads.csv",
    gsheet_id=os.getenv("GOOGLE_SHEET_ID"),
    gservice_json=os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
)
storage = Storage(storage_cfg)

st.set_page_config(page_title=APP_NAME, page_icon="🎯", layout="centered")

# -- LOGO --
if SHOW_LOGO:
    _logo_bytes = load_image_bytes(LOGO_PATH)
    if _logo_bytes:
        st.image(_logo_bytes, width=160)

tab_share, tab_scan, tab_export = st.tabs(["📇 Partager ma carte", "📷 Scanner un QR", "📤 Export"])

with tab_share:
    st.title("Scannez pour ajouter le contact")
    st.caption("Présentez ce QR au visiteur. Sur son smartphone, il ouvrira votre vCard (.vcf).")

    # Image / QR dyn
    photo_bytes: Optional[bytes] = load_image_bytes(PHOTO_PATH)
    qr_bytes: Optional[bytes] = None
    if QR_TARGET_URL:
        try:
            import qrcode
            qr_img = qrcode.make(QR_TARGET_URL)
            buf = io.BytesIO()
            qr_img.save(buf, format="PNG")
            qr_bytes = buf.getvalue()
        except Exception:
            qr_bytes = None
    else:
        qr_bytes = load_image_bytes(QR_IMAGE_PATH)

    if SHOW_QR_IN_HEADER and qr_bytes:
        st.image(qr_bytes, width=300, caption="Scannez le QR pour récupérer ma carte")
    elif photo_bytes:
        st.image(photo_bytes, width=160, caption=IDENTITY["FN"])

    # vCard côté présentateur (optionnel)
    vcard_bytes = build_vcard_bytes(
        fn=IDENTITY["FN"],
        n_last=IDENTITY["N_LAST"],
        n_first=IDENTITY["N_FIRST"],
        org=IDENTITY["ORG"],
        title=IDENTITY["TITLE"],
        tel=IDENTITY["TEL"],
        email=IDENTITY["EMAIL"],
        url=IDENTITY["URL"],
        adr_street=IDENTITY["ADR_STREET"],
        adr_city=IDENTITY["ADR_CITY"],
        adr_pc=IDENTITY["ADR_PC"],
        adr_country=IDENTITY["ADR_COUNTRY"],
        photo_bytes=photo_bytes
    )
    if SHOW_DOWNLOAD_BUTTON:
        st.download_button("📇 Sauvegarder le contact (.vcf)", data=vcard_bytes,
                           file_name=f"{IDENTITY['FN'].replace(' ', '_')}.vcf",
                           mime="text/vcard")

    st.divider()
    st.header("Laissez-moi votre carte ✍️")
    with st.form("lead_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            first_name = st.text_input("Prénom *", max_chars=64)
        with c2:
            last_name = st.text_input("Nom *", max_chars=64)
        c3, c4 = st.columns(2)
        with c3:
            email = st.text_input("Email *", max_chars=120)
        with c4:
            phone = st.text_input("Téléphone", max_chars=32)
        company = st.text_input("Société *", max_chars=120)
        job = st.text_input("Poste", max_chars=120)
        interest = st.selectbox("Intérêt", ["Prise de contact", "Démo", "Devis", "Partenariat", "Autre"])
        consent = st.checkbox("J’accepte d’être recontacté·e (RGPD)")
        submit = st.form_submit_button("Envoyer")

    if submit:
        if not (first_name and last_name and email and company and consent):
            st.error("Merci de remplir les champs obligatoires (*) et d’accepter le contact.")
        else:
            lead = Lead(
                first_name=first_name.strip(),
                last_name=last_name.strip(),
                email=email.strip(),
                phone=phone.strip(),
                company=company.strip(),
                job=job.strip(),
                interest=interest,
                utm_source="form-salon",
                ip_hash=get_client_ip_hash(),
            )
            ok, msg = storage.save_lead(lead)
            if ok:
                st.success("✅ Merci ! Votre carte est bien enregistrée.")
            else:
                st.error(f"Erreur: {msg}")


with tab_scan:
    st.header("Scanner un QR visiteur")
    st.caption("Choisissez la caméra **Arrière**. Si besoin, utilisez le *Mode photo (recommandé)* en dessous.")

    if not QR_ENABLED:
        st.warning("Scanner live indisponible : installez 'streamlit-webrtc', 'opencv-python(-headless)', 'av', 'aiortc', 'numpy'.")
    else:
        import cv2  # type: ignore
        from streamlit_webrtc import webrtc_streamer, WebRtcMode, VideoTransformerBase

        class QRProcessor(VideoTransformerBase):  # type: ignore[misc]
            def __init__(self) -> None:
                self.last_result: Optional[str] = None
                # Reuse detector to reduce per-frame allocations
                self._detector = cv2.QRCodeDetector()
                self._frame_count = 0

            def transform(self, frame):
                img = frame.to_ndarray(format="bgr24")
                self._frame_count += 1

                # Try to decode every 2 frames to lighten CPU
                if self._frame_count % 2 == 0:
                    try:
                        from modules.qr import decode_qr_from_ndarray
                        data = decode_qr_from_ndarray(img)
                        if data:
                            self.last_result = data
                    except Exception:
                        pass

                # Draw a guide when a QR is detected (helps aiming)
                try:
                    ok, points = self._detector.detect(img)
                    if ok and points is not None:
                        for p in points:
                            pts = p.astype(int).reshape((-1, 1, 2))
                            cv2.polylines(img, [pts], True, (0, 255, 0), 2)
                except Exception:
                    pass

                return img
,
            video_transformer_factory=QRProcessor,
            async_processing=True,
        )

        if "qr_last" not in st.session_state:
            st.session_state["qr_last"] = None
        auto_save = st.checkbox("Auto-enregistrer dès qu'un QR valide est détecté", value=False)

        qr_text = ctx.video_transformer.last_result if ctx and ctx.video_transformer else None
        if qr_text:
            with st.expander("Voir le contenu brut du QR détecté"):
                st.code(qr_text, language="text")

            from modules.qr import parse_contact_from_qr
            parsed = parse_contact_from_qr(qr_text)
            st.write("Champs reconnus :", parsed)

            is_new = (qr_text != st.session_state["qr_last"])
            if is_new:
                st.session_state["qr_last"] = qr_text

            with st.form("scan_to_lead", clear_on_submit=False):
                c1, c2 = st.columns(2)
                with c1:
                    s_first = st.text_input("Prénom *", value=parsed.get("first_name",""))
                with c2:
                    s_last = st.text_input("Nom *", value=parsed.get("last_name",""))
                c3, c4 = st.columns(2)
                with c3:
                    s_email = st.text_input("Email *", value=parsed.get("email",""))
                with c4:
                    s_phone = st.text_input("Téléphone", value=parsed.get("phone",""))
                s_company = st.text_input("Société *", value=parsed.get("company",""))
                s_job = st.text_input("Poste", value=parsed.get("job",""))
                s_interest = st.selectbox("Intérêt", ["Prise de contact","Démo","Devis","Partenariat","Autre"])
                s_consent = st.checkbox("J’accepte d’être recontacté·e (RGPD)", value=True)
                s_submit = st.form_submit_button("Enregistrer le lead")

            def _save_lead(first, last, email, phone, company, job, interest, consent):
                if not (first and last and email and company and consent):
                    st.error("Champs obligatoires manquants (Prénom, Nom, Email, Société) ou consentement.")
                    return
                lead = Lead(first_name=first.strip(), last_name=last.strip(),
                            email=email.strip(), phone=phone.strip(),
                            company=company.strip(), job=job.strip(),
                            interest=interest, utm_source="qr-scan",
                            ip_hash=get_client_ip_hash())
                ok, msg = storage.save_lead(lead)
                if ok:
                    st.success("✅ Lead enregistré depuis QR.")
                else:
                    st.error(f"Erreur: {msg}")

            if s_submit:
                _save_lead(s_first, s_last, s_email, s_phone, s_company, s_job, s_interest, s_consent)

            if auto_save and is_new:
                if parsed.get("first_name") and parsed.get("last_name") and parsed.get("email") and parsed.get("company"):
                    _save_lead(parsed["first_name"], parsed["last_name"], parsed["email"],
                               parsed.get("phone",""), parsed["company"], parsed.get("job",""),
                               "Prise de contact", True)
                else:
                    st.info("QR détecté, mais informations incomplètes pour un auto-enregistrement. Complétez puis cliquez sur Enregistrer.")
        else:
            st.info("Aucun QR détecté pour l'instant. Approchez un QR net et bien éclairé à ~15–25 cm.")

    st.subheader("Mode photo (recommandé)")
    st.caption("Utilisez l'appareil photo pour capturer une image nette du QR (très fiable).")
    img_cam = st.camera_input("Cadrez le QR puis prenez la photo")
    if img_cam is not None:
        data = img_cam.getvalue()
        from modules.qr import decode_qr_from_bytes, parse_contact_from_qr
        qr_text2 = decode_qr_from_bytes(data)
        if qr_text2:
            st.success("QR détecté ✔ (mode photo)")
            with st.expander("Voir le contenu brut du QR (photo)"):
                st.code(qr_text2, language="text")
            parsed2 = parse_contact_from_qr(qr_text2)
            with st.form("photo_to_lead", clear_on_submit=False):
                c1, c2 = st.columns(2)
                with c1:
                    p_first = st.text_input("Prénom *", value=parsed2.get("first_name",""))
                with c2:
                    p_last = st.text_input("Nom *", value=parsed2.get("last_name",""))
                c3, c4 = st.columns(2)
                with c3:
                    p_email = st.text_input("Email *", value=parsed2.get("email",""))
                with c4:
                    p_phone = st.text_input("Téléphone", value=parsed2.get("phone",""))
                p_company = st.text_input("Société *", value=parsed2.get("company",""))
                p_job = st.text_input("Poste", value=parsed2.get("job",""))
                p_interest = st.selectbox("Intérêt", ["Prise de contact","Démo","Devis","Partenariat","Autre"])
                p_consent = st.checkbox("J’accepte d’être recontacté·e (RGPD)", value=True)
                p_submit = st.form_submit_button("Enregistrer le lead")
            if p_submit:
                if not (p_first and p_last and p_email and p_company and p_consent):
                    st.error("Champs obligatoires manquants ou consentement non coché.")
                else:
                    lead = Lead(first_name=p_first.strip(), last_name=p_last.strip(),
                                email=p_email.strip(), phone=p_phone.strip(),
                                company=p_company.strip(), job=p_job.strip(),
                                interest=p_interest, utm_source="qr-photo",
                                ip_hash=get_client_ip_hash())
                    ok, msg = storage.save_lead(lead)
                    if ok:
                        st.success("✅ Lead enregistré depuis photo.")
                    else:
                        st.error(f"Erreur: {msg}")

    st.subheader("Ou importer une image de QR")
    img = st.file_uploader("Photo du QR (PNG/JPG)", type=["png","jpg","jpeg"])
    if img is not None:
        from modules.qr import decode_qr_from_bytes
        data = decode_qr_from_bytes(img.read())
        if data:
            st.info("QR décodé 👍")
            st.code(data, language="text")
        else:
            st.error("Impossible de décoder un QR dans cette image.")

with tab_export:
    st.header("Exporter les leads")
    st.caption("CSV simple ou Excel (.xlsx) au format tableau avec filtres.")
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows

    csv_path = DATA_DIR / "leads.csv"
    if not csv_path.exists():
        st.info("Aucun lead pour l’instant.")
    else:
        try:
            df = pd.read_csv(csv_path, dtype=str).fillna("")
            st.dataframe(df, use_container_width=True)
            st.download_button("📄 Télécharger CSV", data=csv_path.read_bytes(),
                               file_name="leads.csv", mime="text/csv")
            # Excel stylé
            from io import BytesIO
            wb = Workbook()
            ws = wb.active; ws.title = "Leads"
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            # Délimitation de la table
            last_col = ws.max_column
            last_row = ws.max_row
            # Générer une référence de table dynamique (colonnes A..Z max)
            from string import ascii_uppercase
            col_letter = ascii_uppercase[last_col-1] if last_col <= 26 else "Z"
            ref = f"A1:{col_letter}{last_row}"
            table = Table(displayName="LeadsTable", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9",
                                   showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
            ws.freeze_panes = "A2"
            # Largeurs
            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for c in col:
                    max_len = max(max_len, len(str(c.value)) if c.value else 0)
                ws.column_dimensions[letter].width = min(max(12, max_len + 2), 40)
            buf = BytesIO(); wb.save(buf)
            st.download_button("📊 Télécharger Excel (.xlsx)", data=buf.getvalue(),
                               file_name="leads.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erreur export: {e}")
