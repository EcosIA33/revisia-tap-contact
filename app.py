#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Salon edition:
- QR en en-tête (image ou généré dynamiquement via QR_TARGET_URL)
- Bouton .vcf masquable (SHOW_DOWNLOAD_BUTTON)
- Scanner QR (caméra arrière par défaut)
- Export CSV + Excel (.xlsx) avec vrai tableau stylé
"""
from __future__ import annotations
import os
import io
import logging
from pathlib import Path
from typing import Optional, Dict

import streamlit as st
from dotenv import load_dotenv

from modules.contact import build_vcard_bytes
from modules.storage import Lead, Storage, StorageConfig
from modules.utils import ensure_data_dir, get_client_ip_hash, load_image_bytes
from modules.qr import parse_contact_from_qr

# WebRTC imports (optionnels)
try:
    import av  # noqa: F401
    import cv2  # noqa: F401
    from streamlit_webrtc import webrtc_streamer, WebRtcMode, VideoTransformerBase
    QR_ENABLED = True
except Exception:
    QR_ENABLED = False

APP_NAME = "Tap-to-Contact — Salon"

# --- Boot ---
load_dotenv()
DATA_DIR = Path(os.getenv("DATA_DIR", "data"))
LOGS_DIR = Path(os.getenv("LOGS_DIR", "logs"))
ensure_data_dir(DATA_DIR)

# --- Identité (.env) ---
IDENTITY: Dict[str, str] = {
    "FN": os.getenv("FULL_NAME", "Votre Nom"),
    "N_LAST": os.getenv("N_LAST", "Nom"),
    "N_FIRST": os.getenv("N_FIRST", "Prénom"),
    "ORG": os.getenv("ORG", "Votre Société"),
    "TITLE": os.getenv("TITLE", "Fonction"),
    "TEL": os.getenv("TEL", "+33 6 00 00 00 00"),
    "EMAIL": os.getenv("EMAIL", "vous@example.com"),
    "URL": os.getenv("URL", "https://www.exemple.com"),
    "ADR_STREET": os.getenv("ADR_STREET", ""),
    "ADR_CITY": os.getenv("ADR_CITY", ""),
    "ADR_PC": os.getenv("ADR_PC", ""),
    "ADR_COUNTRY": os.getenv("ADR_COUNTRY", "France"),
}

PHOTO_PATH = os.getenv("PHOTO_PATH", "assets/photo.jpg")
SHOW_QR_IN_HEADER = os.getenv("SHOW_QR_IN_HEADER", "true").lower() in ("1","true","yes","on")
QR_IMAGE_PATH = os.getenv("QR_IMAGE_PATH", "assets/qr.png")
QR_TARGET_URL = os.getenv("QR_TARGET_URL", "")  # si défini, on génère le QR dynamiquement
SHOW_DOWNLOAD_BUTTON = os.getenv("SHOW_DOWNLOAD_BUTTON", "false").lower() in ("1","true","yes","on")

# --- Storage ---
storage_cfg = StorageConfig(
    csv_path=DATA_DIR / "leads.csv",
    gsheet_id=os.getenv("GOOGLE_SHEET_ID"),
    gservice_json=os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
)
storage = Storage(storage_cfg)

st.set_page_config(page_title=APP_NAME, page_icon="🎯", layout="centered")

tab_share, tab_scan, tab_export = st.tabs(["📇 Partager ma carte", "📷 Scanner un QR", "📤 Export"])

with tab_share:
    st.title("Scannez pour ajouter le contact")
    st.caption("Présentez ce QR au visiteur. Sur son smartphone, il ouvrira votre vCard (.vcf).")

    # Image / QR dyn
    photo_bytes: Optional[bytes] = load_image_bytes(PHOTO_PATH)
    qr_bytes: Optional[bytes] = None
    if QR_TARGET_URL:
        try:
            import qrcode
            qr_img = qrcode.make(QR_TARGET_URL)
            buf = io.BytesIO()
            qr_img.save(buf, format="PNG")
            qr_bytes = buf.getvalue()
        except Exception:
            qr_bytes = None
    else:
        qr_bytes = load_image_bytes(QR_IMAGE_PATH)

    if SHOW_QR_IN_HEADER and qr_bytes:
        st.image(qr_bytes, width=300, caption="Scannez le QR pour récupérer ma carte")
    elif photo_bytes:
        st.image(photo_bytes, width=160, caption=IDENTITY["FN"])

    # vCard côté présentateur (optionnel)
    vcard_bytes = build_vcard_bytes(
        fn=IDENTITY["FN"],
        n_last=IDENTITY["N_LAST"],
        n_first=IDENTITY["N_FIRST"],
        org=IDENTITY["ORG"],
        title=IDENTITY["TITLE"],
        tel=IDENTITY["TEL"],
        email=IDENTITY["EMAIL"],
        url=IDENTITY["URL"],
        adr_street=IDENTITY["ADR_STREET"],
        adr_city=IDENTITY["ADR_CITY"],
        adr_pc=IDENTITY["ADR_PC"],
        adr_country=IDENTITY["ADR_COUNTRY"],
        photo_bytes=photo_bytes
    )
    if SHOW_DOWNLOAD_BUTTON:
        st.download_button("📇 Sauvegarder le contact (.vcf)", data=vcard_bytes,
                           file_name=f"{IDENTITY['FN'].replace(' ', '_')}.vcf",
                           mime="text/vcard")

    st.divider()
    st.header("Laissez-moi votre carte ✍️")
    with st.form("lead_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            first_name = st.text_input("Prénom *", max_chars=64)
        with c2:
            last_name = st.text_input("Nom *", max_chars=64)
        c3, c4 = st.columns(2)
        with c3:
            email = st.text_input("Email *", max_chars=120)
        with c4:
            phone = st.text_input("Téléphone", max_chars=32)
        company = st.text_input("Société *", max_chars=120)
        job = st.text_input("Poste", max_chars=120)
        interest = st.selectbox("Intérêt", ["Prise de contact", "Démo", "Devis", "Partenariat", "Autre"])
        consent = st.checkbox("J’accepte d’être recontacté·e (RGPD)")
        submit = st.form_submit_button("Envoyer")

    if submit:
        if not (first_name and last_name and email and company and consent):
            st.error("Merci de remplir les champs obligatoires (*) et d’accepter le contact.")
        else:
            lead = Lead(
                first_name=first_name.strip(),
                last_name=last_name.strip(),
                email=email.strip(),
                phone=phone.strip(),
                company=company.strip(),
                job=job.strip(),
                interest=interest,
                utm_source="form-salon",
                ip_hash=get_client_ip_hash(),
            )
            ok, msg = storage.save_lead(lead)
            if ok:
                st.success("✅ Merci ! Votre carte est bien enregistrée.")
            else:
                st.error(f"Erreur: {msg}")

with tab_scan:
    st.header("Scanner un QR visiteur")
    st.caption("La caméra arrière est utilisée par défaut. Si besoin, changez de caméra dans le menu du navigateur.")
    if not QR_ENABLED:
        st.warning("Scanner live indisponible : installez 'streamlit-webrtc', 'opencv-python(-headless)', 'av', 'aiortc', 'numpy'.")
    else:
        import cv2  # type: ignore
        from streamlit_webrtc import webrtc_streamer, WebRtcMode, VideoTransformerBase

        class QRProcessor(VideoTransformerBase):  # type: ignore[misc]
            def __init__(self) -> None:
                self.detector = cv2.QRCodeDetector()
                self.last_result: Optional[str] = None
            def transform(self, frame):
                img = frame.to_ndarray(format="bgr24")
                try:
                    data, points, _ = self.detector.detectAndDecode(img)
                    if data:
                        self.last_result = data.strip()
                except Exception:
                    pass
                return img

        ctx = webrtc_streamer(
            key="qr-scan",
            mode=WebRtcMode.SENDRECV,
            media_stream_constraints={"video": {"facingMode": {"ideal": "environment"}}, "audio": False},
            video_transformer_factory=QRProcessor,
            async_processing=True,
        )

        qr_text = ctx.video_transformer.last_result if ctx and ctx.video_transformer else None
        if qr_text:
            st.success("QR détecté ✔")
            from modules.qr import parse_contact_from_qr
            parsed = parse_contact_from_qr(qr_text)
            with st.form("scan_to_lead", clear_on_submit=False):
                c1, c2 = st.columns(2)
                with c1:
                    s_first = st.text_input("Prénom *", value=parsed.get("first_name",""))
                with c2:
                    s_last = st.text_input("Nom *", value=parsed.get("last_name",""))
                c3, c4 = st.columns(2)
                with c3:
                    s_email = st.text_input("Email *", value=parsed.get("email",""))
                with c4:
                    s_phone = st.text_input("Téléphone", value=parsed.get("phone",""))
                s_company = st.text_input("Société *", value=parsed.get("company",""))
                s_job = st.text_input("Poste", value=parsed.get("job",""))
                s_interest = st.selectbox("Intérêt", ["Prise de contact","Démo","Devis","Partenariat","Autre"])
                s_consent = st.checkbox("J’accepte d’être recontacté·e (RGPD)", value=True)
                s_submit = st.form_submit_button("Enregistrer le lead")
            if s_submit:
                if not (s_first and s_last and s_email and s_company and s_consent):
                    st.error("Champs obligatoires manquants ou consentement non coché.")
                else:
                    lead = Lead(
                        first_name=s_first.strip(),
                        last_name=s_last.strip(),
                        email=s_email.strip(),
                        phone=s_phone.strip(),
                        company=s_company.strip(),
                        job=s_job.strip(),
                        interest=s_interest,
                        utm_source="qr-scan",
                        ip_hash=get_client_ip_hash(),
                    )
                    ok, msg = storage.save_lead(lead)
                    if ok:
                        st.success("✅ Lead enregistré depuis QR.")
                    else:
                        st.error(f"Erreur: {msg}")

    st.subheader("Ou importer une image de QR")
    img = st.file_uploader("Photo du QR (PNG/JPG)", type=["png","jpg","jpeg"])
    if img is not None:
        from modules.qr import decode_qr_from_bytes
        data = decode_qr_from_bytes(img.read())
        if data:
            st.info("QR décodé 👍")
            st.code(data, language="text")
        else:
            st.error("Impossible de décoder un QR dans cette image.")

with tab_export:
    st.header("Exporter les leads")
    st.caption("CSV simple ou Excel (.xlsx) au format tableau avec filtres.")
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils.dataframe import dataframe_to_rows

    csv_path = DATA_DIR / "leads.csv"
    if not csv_path.exists():
        st.info("Aucun lead pour l’instant.")
    else:
        try:
            df = pd.read_csv(csv_path, dtype=str).fillna("")
            st.dataframe(df, use_container_width=True)
            st.download_button("📄 Télécharger CSV", data=csv_path.read_bytes(),
                               file_name="leads.csv", mime="text/csv")
            # Excel stylé
            from io import BytesIO
            wb = Workbook()
            ws = wb.active; ws.title = "Leads"
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            # Délimitation de la table
            last_col = ws.max_column
            last_row = ws.max_row
            # Générer une référence de table dynamique (colonnes A..Z max)
            from string import ascii_uppercase
            col_letter = ascii_uppercase[last_col-1] if last_col <= 26 else "Z"
            ref = f"A1:{col_letter}{last_row}"
            table = Table(displayName="LeadsTable", ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9",
                                   showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
            ws.freeze_panes = "A2"
            # Largeurs
            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for c in col:
                    max_len = max(max_len, len(str(c.value)) if c.value else 0)
                ws.column_dimensions[letter].width = min(max(12, max_len + 2), 40)
            buf = BytesIO(); wb.save(buf)
            st.download_button("📊 Télécharger Excel (.xlsx)", data=buf.getvalue(),
                               file_name="leads.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erreur export: {e}")
